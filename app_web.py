import os
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import time

# --- CONFIGURATION ---
FICHIERS = {"CARACT ENTRANT": "Suivi CARACT_ENTRANT.xlsx", "CARACT SORTANT": "CARACT_SORTANT.xlsx"}
LISTE_EQUIPES = ["MATIN", "APRES-MIDI"]
LISTE_LIEUX = ["SILO", "CABINE", "COMPACTEUR", "COLLECTE"]
NOM_LOGO = "PAPREC_Logotype_V_BLACK.jpg"

COLORS = {"BLUE": "#0070c0", "LIGHT_BLUE": "#00b0f0", "GREEN": "#00b050", "YELLOW": "#ffc000", "ORANGE": "#ed7d31", "RED": "#c00000"}

# --- LISTES CLIENTS ET FLUX ---
LISTE_CLIENTS_ENTRANT = ["GARE MONTPARNASSE", "LE PETIT PLUS", "LA COURNEUVE", "PSM", "PSM - ADP BEAUVAIS", "LBM", "AEROVILLE", "CDG", "HAUSSMAN", "ORLY", "CEMEX", "LE BOURGET & LBM", "ROISSY", "DISNEYLAND", "VALLEE VILLAGE & 4 TEMPS", "GARE DE LYON", "SNCF", "VALODEA", "VALOR'AISME", "SITRU", "SYCTOM", "AUTRES"]
LISTE_FLUX_SORTANT = ["JRM", "EMR", "CARTON", "GM", "PETQ9", "FLUX DEV", "PETB", "PEPP", "ELA", "FILM", "ACIER", "ALU", "PETIT ALU", "REFUS"]

GROUPES_ENTRANT = {
    "FIBREUX": {"color": COLORS["BLUE"], "items": ["CARTON", "CARTONNETTE"]},
    "FIBREUX 2": {"color": COLORS["LIGHT_BLUE"], "items": ["ECRIT COULEUR", "JRM", "GM"]},
    "ELA": {"color": COLORS["GREEN"], "items": ["ELA"]},
    "PLASTIQUES": {"color": COLORS["YELLOW"], "items": ["PEPP", "PET Q9", "PET Q5", "PET B", "FILM"]},
    "METAUX": {"color": COLORS["ORANGE"], "items": ["ACIER", "ALU", "PETIT ALU"]},
    "REFUS / AUTRES": {"color": COLORS["RED"], "items": ["PB REFUSES", "FILM REFUSES", "EMBALLAGES NOIRS", "AUTRES EMBALLAGES NON RECYCLABLES", "BOIS", "VERRE", "DDS", "D3E", "IMBRIQUES", "PLASTIQUES NON EMBALLAGES", "FERRAILLES", "GRAVATS", "TEXTILE", "EMBALLAGES NON VIDES / SOUILLÉS", "REFUS", "PAPIER MOUILLE", "FINES"]}
}

GROUPES_SORTANT = {
    "FIBREUX": {"color": COLORS["BLUE"], "items": ["JRM", "EMR", "CARTON", "GM"]},
    "PLASTIQUES": {"color": COLORS["YELLOW"], "items": ["PETQ9", "FLUX DEV", "PET B", "PEPP", "ELA", "FILM"]},
    "METAUX": {"color": COLORS["ORANGE"], "items": ["ACIER", "ALU", "PETIT ALU"]},
    "AUTRES": {"color": COLORS["RED"], "items": ["REFUS"]}
}

# --- FONCTIONS ---
def afficher_logo(largeur=None):
    if os.path.exists(NOM_LOGO):
        if largeur: st.image(NOM_LOGO, width=largeur)
        else: st.image(NOM_LOGO, use_container_width=True)
    else: st.write("### PAPREC")

def enregistrer_donnees(mode, header, dict_poids):
    nom_f = FICHIERS[mode]
    if not os.path.exists(nom_f):
        st.error(f"Fichier {nom_f} introuvable.")
        return False
    try:
        wb = load_workbook(nom_f)
        ws = wb["SAISIE"]
        row = 2
        while ws.cell(row=row, column=2).value: row += 1
        groupes = GROUPES_ENTRANT if mode == "CARACT ENTRANT" else GROUPES_SORTANT
        liste_ordonnee = []
        for g in groupes.values(): liste_ordonnee.extend(g["items"])
        poids_finaux = [float(str(dict_poids.get(m, 0)).replace(',', '.')) for m in liste_ordonnee]
        
        if mode == "CARACT SORTANT":
            ligne = [header['flux'], header['date'], header['equipe'], header['lieu']] + poids_finaux
        else:
            ligne = [header['flux'], header['date']] + poids_finaux
            
        for i, v in enumerate(ligne):
            cell = ws.cell(row=row, column=i+2, value=v)
            cell.alignment = Alignment(horizontal="center")
        wb.save(nom_f)
        return True
    except Exception as e:
        st.error(f"❌ Erreur Excel : {e}")
        return False

# --- CONFIGURATION PAGE ---
st.set_page_config(page_title="PAPREC - Caractérisation", layout="wide")

if 'mode' not in st.session_state: st.session_state.mode = None
if 'photos_temp' not in st.session_state: st.session_state.photos_temp = {}

# --- ECRAN D'ACCUEIL ---
if st.session_state.mode is None:
    col_l1, col_l2, col_l3 = st.columns([1, 1.2, 1])
    with col_l2: afficher_logo()
    st.markdown("<h1 style='text-align: center; color: #0070c0;'>Fiche de caractérisation</h1>", unsafe_allow_html=True)
    st.write("##") 
    c1, c2 = st.columns(2)
    if c1.button("📥 CARACT ENTRANT", use_container_width=True):
        st.session_state.mode = "CARACT ENTRANT"
        st.rerun()
    if c2.button("📤 CARACT SORTANT", use_container_width=True):
        st.session_state.mode = "CARACT SORTANT"
        st.rerun()

# --- ECRAN DE SAISIE ---
else:
    col_back, col_logo_mini = st.columns([8, 2])
    with col_back:
        if st.button("⬅ Retour"):
            st.session_state.mode = None
            st.session_state.photos_temp.clear()
            st.rerun()
    with col_logo_mini: afficher_logo(largeur=100)

    st.markdown(f"<h2 style='text-align: center;'>Saisie : {st.session_state.mode}</h2>", unsafe_allow_html=True)

    with st.container(border=True):
        col1, col2, col3, col4 = st.columns(4)
        date_saisie = col1.text_input("Date (JJ/MM/AAAA):", datetime.now().strftime("%d/%m/%Y"))
        lbl = "Client:" if st.session_state.mode == "CARACT ENTRANT" else "Flux:"
        lst = LISTE_CLIENTS_ENTRANT if st.session_state.mode == "CARACT ENTRANT" else LISTE_FLUX_SORTANT
        flux_sel = col2.selectbox(lbl, lst)
        equipe_sel = col3.selectbox("Équipe:", LISTE_EQUIPES)
        lieu_sel = col4.selectbox("Lieu:", LISTE_LIEUX)

    groupes = GROUPES_ENTRANT if st.session_state.mode == "CARACT ENTRANT" else GROUPES_SORTANT
    dict_entrees = {}
    
    for g_name, info in groupes.items():
        st.markdown(f"<div style='background-color:{info['color']}; color:white; padding:10px; border-radius:5px; margin-top:20px; margin-bottom:10px;'><b>{g_name}</b></div>", unsafe_allow_html=True)
        items = info["items"]
        for i in range(0, len(items), 2):
            cols = st.columns(2)
            for j in range(2):
                if i + j < len(items):
                    matiere = items[i+j]
                    with cols[j].container(border=True):
                        c_poids, c_photo = st.columns([1, 1])
                        # Initialisation du poids dans les entrées
                        dict_entrees[matiere] = c_poids.number_input(f"{matiere} (kg)", min_value=0.0, step=0.1, key=f"p_{matiere}")
                        
                        with c_photo.popover("📸 Photo"):
                            if matiere in st.session_state.photos_temp:
                                st.image(st.session_state.photos_temp[matiere], caption="Photo enregistrée", width=200)
                                if st.button(f"🗑️ Supprimer la photo", key=f"del_{matiere}"):
                                    del st.session_state.photos_temp[matiere]
                                    st.rerun()
                            else:
                                img = st.camera_input(f"Capturer {matiere}", key=f"cam_{matiere}")
                                if img:
                                    st.image(img, caption="Aperçu", width=200)
                                    col_v1, col_v2 = st.columns(2)
                                    if col_v1.button("✅ Sauvegarder", key=f"save_{matiere}", type="primary"):
                                        st.session_state.photos_temp[matiere] = img
                                        st.rerun()
                                    if col_v2.button("❌ Annuler", key=f"cancel_{matiere}"):
                                        st.rerun()

    st.markdown("---")
    # --- ACTION FINALE D'ENREGISTREMENT ---
    if st.button("💾 ENREGISTRER DÉFINITIVEMENT (EXCEL + PHOTOS)", type="primary", use_container_width=True):
        date_folder = date_saisie.replace("/", "-")
        nom_dossier = f"{flux_sel}_{date_folder}"
        sous_type = "ENTRANT" if st.session_state.mode == "CARACT ENTRANT" else "SORTANT"
        path_complet = os.path.join("PHOTOS_CARACT", sous_type, nom_dossier)
        
        try:
            # 1. Sauvegarde des photos
            if st.session_state.photos_temp:
                if not os.path.exists(path_complet): os.makedirs(path_complet)
                for mat, data in st.session_state.photos_temp.items():
                    with open(os.path.join(path_complet, f"{mat}.jpg"), "wb") as f:
                        f.write(data.getbuffer())

            # 2. Sauvegarde Excel
            h = {'date': date_saisie, 'flux': flux_sel, 'equipe': equipe_sel, 'lieu': lieu_sel}
            if enregistrer_donnees(st.session_state.mode, h, dict_entrees):
                # 3. Message de succès et redirection
                st.success(f"✅ Enregistrement réussi ! Retour à l'accueil...")
                st.balloons()
                
                # Attente courte pour que l'utilisateur voit le message
                time.sleep(2)
                
                # 4. RÉINITIALISATION COMPLÈTE
                st.session_state.mode = None
                st.session_state.photos_temp.clear()
                
                # Nettoyage des widgets (poids à zéro)
                for key in list(st.session_state.keys()):
                    if key.startswith("p_") or key.startswith("cam_"):
                        del st.session_state[key]
                
                st.rerun()
                
        except Exception as e:
            st.error(f"❌ Erreur lors de la sauvegarde : {e}")
